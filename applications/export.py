import io
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from .models import ApplicationTranscript

def get_matched_results_for_application(application):
    results = []
    prereq_maps = application.prerequisitemap_set.prefetch_related('prerequisitemapcourses_set__course').all()
    
    transcripts = {t.course_id: t for t in ApplicationTranscript.objects.filter(application=application)}

    for pmap in prereq_maps:
        taken = []
        for map_course in pmap.prerequisitemapcourses_set.all():
            course = map_course.course
            transcript = transcripts.get(course.course_id)
            taken.append({
                "course": course,
                "grade": transcript.grade if transcript else "N/A"
            })
        
        results.append({
            "target_course": pmap.target_course,
            "taken_courses": taken
        })
    return results

def generate_csv_for_application(application):
    matched_results = get_matched_results_for_application(application)
    
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)

    writer.writerow(["PREREQS"])

    if matched_results:
        target_codes = [r["target_course"].course_code if r["target_course"] else "N/A" for r in matched_results]
        target_names = [r["target_course"].course_name if r["target_course"] else "N/A" for r in matched_results]

        writer.writerow(target_codes)
        writer.writerow(target_names)

        max_len = max([len(r["taken_courses"]) for r in matched_results] + [0])
        
        for i in range(max_len):
            row = []
            for r in matched_results:
                if i < len(r["taken_courses"]):
                    taken = r["taken_courses"][i]
                    course = taken["course"]
                    grade = taken["grade"]
                    row.append(f"{course.course_code} - {course.course_name} (Grade: {grade})")
                else:
                    row.append("")
            writer.writerow(row)

    buffer.seek(0)
    return buffer

def generate_xlsx_for_application(application):
    matched_results = get_matched_results_for_application(application)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Courses"

    header_fill = PatternFill(start_color='B7D7F7', end_color='B7D7F7', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_align = Alignment(wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)

    target_codes = [r["target_course"].course_code if r["target_course"] else "N/A" for r in matched_results] if matched_results else []
    target_names = [r["target_course"].course_name if r["target_course"] else "N/A" for r in matched_results] if matched_results else []
    num_columns = len(target_codes) if target_codes else 1

    # Row 1: "PREREQS" merged
    ws.append(["PREREQS"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    cell = ws.cell(row=1, column=1)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    cell.font = bold_font

    if matched_results:
        ws.append(target_codes)
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = bold_font

        ws.append(target_names)
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = bold_font

        max_len = max([len(r["taken_courses"]) for r in matched_results] + [0])
        for i in range(max_len):
            row = []
            for r in matched_results:
                if i < len(r["taken_courses"]):
                    taken = r["taken_courses"][i]
                    course = taken["course"]
                    grade = taken["grade"]
                    row.append(f"{course.course_code} - {course.course_name} (Grade: {grade})")
                else:
                    row.append("")
            ws.append(row)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=num_columns):
        for cell in row:
            if cell.value is not None:
                cell.alignment = wrap_align
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 5, 50)

    return wb
