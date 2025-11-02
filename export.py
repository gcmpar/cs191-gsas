import io
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

def generate_csv_for_student(student_id, matched_results):
    buffer = io.StringIO()  # Use StringIO for text data
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)

    writer.writerow(["PREREQS"])

    if matched_results:
        core_courses = [r.get("core_course_code", "") for r in matched_results]
        prereq_codes = [r["prereq_course_code"] for r in matched_results]

        writer.writerow(core_courses)
        writer.writerow(prereq_codes)

        # Determine the max length for matching courses in the row
        max_len = max(len(r["matched_courses"]) for r in matched_results)
        


        # Loop through and write each match for the courses
        for i in range(max_len):
            row = []
            for r in matched_results:
                if i < len(r["matched_courses"]):
                    match = r["matched_courses"][i]
                    row.append(f"{match['taken_course_code']} - {match['taken_description']} (Grade: {match['grade']})")
                else:
                    row.append("")  # Ensure there's a placeholder for missing matches
            writer.writerow(row)

    buffer.seek(0)  # Reset pointer to the beginning before returning the buffer
    return buffer  # Return the StringIO buffer directly

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

def generate_xlsx_for_student(student_id, matched_results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Courses"

    # Styling
    header_fill = PatternFill(start_color='B7D7F7', end_color='B7D7F7', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_align = Alignment(wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)

    # Prepare data
    core_courses = [r.get("core_course_code", "") for r in matched_results] if matched_results else []
    prereq_codes = [r["prereq_course_code"] for r in matched_results] if matched_results else []
    num_columns = len(core_courses)

    # Row 1: "PREREQS" merged
    ws.append(["PREREQS"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns + 1)
    cell = ws.cell(row=1, column=1)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    cell.font = bold_font

    # Row 2: Core course codes
    ws.append(core_courses)
    for col in range(1, num_columns + 2):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = bold_font

    # Row 3: Prerequisite codes
    ws.append(prereq_codes)
    for col in range(1, num_columns + 2):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = bold_font

    # Matched courses
    if matched_results:
        max_len = max(len(r["matched_courses"]) for r in matched_results)
        for i in range(max_len):
            row = []
            for r in matched_results:
                if i < len(r["matched_courses"]):
                    match = r["matched_courses"][i]
                    grade = match.get("grade", "N/A")
                    row.append(f"{match['taken_course_code']} - {match['taken_description']} (Grade: {grade})")
                else:
                    row.append("")
            ws.append([""] + row)

    # Apply border and wrap to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=num_columns + 1):
        for cell in row:
            if cell.value is not None:
                cell.alignment = wrap_align
            cell.border = thin_border

    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 5, 50)

    return wb
